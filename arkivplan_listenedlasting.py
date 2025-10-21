import json
import requests
import os
from openpyxl import Workbook

# === CONFIGURATION ===
site_url = "https://tromkom.sharepoint.com/sites/O365-ProsjektervedArkiv-BK-plan"
auth_json = "auth.json"
output_folder = "lists_excel"
max_items = 1000

# === STEP 1: Extract Cookies from auth.json ===
def get_auth_headers():
    with open(auth_json, "r") as f:
        state = json.load(f)
    cookies = {c["name"]: c["value"] for c in state["cookies"]}
    headers = {
        "Cookie": "; ".join([f"{k}={v}" for k, v in cookies.items()]),
        "Accept": "application/json;odata=verbose"
    }
    return headers

# === STEP 2: Get All Lists and Filter Out System Lists ===
def get_user_lists():
    headers = get_auth_headers()
    endpoint = f"{site_url}/_api/web/lists?$select=Title,Hidden"
    response = requests.get(endpoint, headers=headers)
    if response.status_code != 200:
        print(f"Klarte ikke å hente lister: {response.status_code}")
        return []
    data = response.json()
    all_lists = data.get("d", {}).get("results", [])
    # Filter out hidden/system lists
    excluded_names = {"TaxonomyHiddenList", "appdata", "appfiles"}
    user_lists = [lst for lst in all_lists if not lst["Hidden"] and lst["Title"] not in excluded_names]
    return user_lists

# === STEP 3: Download Items from Each List and Export to Excel ===
def save_list_to_excel(list_title):
    headers = get_auth_headers()
    endpoint = f"{site_url}/_api/web/lists/getbytitle('{list_title}')/items?$top={max_items}"
    response = requests.get(endpoint, headers=headers)
    if response.status_code != 200:
        print(f"Klarte ikke å hente lister fra {list_title}: {response.status_code}")
        return
    items = response.json().get("d", {}).get("results", [])
    if not items:
        print(f"Ingen elementer i '{list_title}'. Listen lagres ikke.")
        return

    # Filter fields: exclude system fields and complex objects
    def clean_item(item):
        return {
            k: (v.get("uri") if isinstance(v, dict) and "__deferred" in v else v)
            for k, v in item.items()
            if not k.startswith("__") and isinstance(v, (str, int, float, bool)) or (isinstance(v, dict) and "__deferred" in v)
        }

    cleaned_items = [clean_item(item) for item in items]
    headers = sorted({key for item in cleaned_items for key in item.keys()})

    wb = Workbook()
    ws = wb.active
    ws.title = list_title[:31]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Write rows
    for row_num, item in enumerate(cleaned_items, 2):
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=row_num, column=col_num, value=item.get(header))

    os.makedirs(output_folder, exist_ok=True)
    wb.save(os.path.join(output_folder, f"{list_title}.xlsx"))
    print(f"✓ Lagret liste '{list_title}' i Excel")

# === MAIN EXECUTION ===
if __name__ == "__main__":
    lists = get_user_lists()
    for lst in lists:
        title = lst["Title"]
        print(f"Laster ned {title}")
        save_list_to_excel(title)